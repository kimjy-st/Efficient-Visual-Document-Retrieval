import os
import re
import json
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from statistics import mean
# -------------------------
# 0) 유틸: 이름 정규화 / mf 추출
# -------------------------
def normalize_dataset_name(raw: str) -> str:
    s = (raw or "").strip().lower()
    if not s:
        return s
    if s == "arxivqa":
        return "arxiv"
    return s

RE_MF_DIR = re.compile(r"^mf\s*([0-9]+(?:\.[0-9]+)?)$", re.IGNORECASE)

def extract_mf_from_dirname(name: str):
    m = RE_MF_DIR.match(name.strip())
    if not m:
        return None
    v = float(m.group(1))
    return int(v) if abs(v - int(v)) < 1e-9 else v

# -------------------------
# 1) train.log 파서: best_ndcg5 기준의 (NDCG@5, Recall@1)
# -------------------------
# JSON line 예: ... {"summary/best_ndcg5": {...}} ...
RE_SUMMARY_JSON = re.compile(r"(\{.*\"summary\/best_ndcg5\".*\})\s*$")

def parse_train_log_best_metrics(train_log_path: str):
    """
    반환: {"N@5": float, "R@1": float} 또는 None
    - train.log에서 summary/best_ndcg5 라인을 뒤에서부터 찾아
      그 안의 {"NDCG@5": x, "Recall@1": y} 추출
    """
    if not os.path.exists(train_log_path):
        return None

    try:
        with open(train_log_path, "r", encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()
    except Exception:
        return None

    tail = lines[-3000:]

    for line in reversed(tail):
        s = line.strip()
        if not s:
            continue

        m = RE_SUMMARY_JSON.search(s)
        if not m:
            continue

        try:
            obj = json.loads(m.group(1))
        except Exception:
            continue

        d = obj.get("summary/best_ndcg5")
        if not isinstance(d, dict):
            continue

        ndcg = d.get("NDCG@5")
        recall = d.get("Recall@1")
        if ndcg is None or recall is None:
            continue

        try:
            ndcg = float(ndcg)
            recall = float(recall)

            # 로그가 0~1이면 표에서는 보통 %로 보니까 100배
            if 0.0 <= ndcg <= 1.0:
                ndcg *= 100.0
            if 0.0 <= recall <= 1.0:
                recall *= 100.0

            return {"N@5": ndcg, "R@1": recall}
        except Exception:
            return None

    return None

# -------------------------
# 2) 엑셀 스타일 유틸
# -------------------------
def autosize_columns(ws):
    max_width = 28
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

def style_header(ws, header_row=1):
    bold = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[header_row]:
        cell.font = bold
        cell.alignment = align

def add_table(ws, last_row, last_col):
    ref = f"A1:{get_column_letter(last_col)}{last_row}"
    tab = Table(displayName=f"Table_{ws.title}".replace(".", "_"), ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

def write_sheet(ws, setting_map, datasets_ordered):
    header = ["setting"]
    for ds in datasets_ordered:
        header += [f"{ds}_N@5", f"{ds}_R@1"]
    header += ["Avg_N@5", "Avg_R@1"]
    ws.append(header)

    for setting in sorted(setting_map.keys()):
        per_ds = setting_map[setting]
        row = [setting]

        n5_list = []
        r1_list = []

        for ds in datasets_ordered:
            if ds in per_ds:
                n5 = round(per_ds[ds]["N@5"], 1)
                r1 = round(per_ds[ds]["R@1"], 1)
                row += [n5, r1]
                n5_list.append(n5)
                r1_list.append(r1)
            else:
                row += [None, None]

        row += [
            round(mean(n5_list), 1) if n5_list else None,
            round(mean(r1_list), 1) if r1_list else None,
        ]
        ws.append(row)

    style_header(ws, 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for r in range(2, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            cell.number_format = "0.0"

    autosize_columns(ws)
    add_table(ws, ws.max_row, ws.max_column)

# -------------------------
# 3) 메인: root_dir 구조 스캔
# -------------------------
def build_excel_from_trainlogs(root_dir: str, out_xlsx: str):
    root_dir = os.path.abspath(root_dir)
    os.makedirs(os.path.dirname(os.path.abspath(out_xlsx)), exist_ok=True)

    settings = [
        d for d in os.listdir(root_dir)
        if os.path.isdir(os.path.join(root_dir, d)) and not d.startswith(".")
    ]
    if not settings:
        raise RuntimeError(f"No setting directories found under: {root_dir}")

    mf_data = defaultdict(lambda: defaultdict(dict))  # mf -> setting -> ds -> metrics
    all_datasets_by_mf = defaultdict(set)

    miss_print_budget = 10  # 파싱 실패 로그 너무 많지 않게

    for setting in sorted(settings):
        setting_path = os.path.join(root_dir, setting)
        setting_key = setting  # 폴더명 그대로

        mf_dirs = [
            d for d in os.listdir(setting_path)
            if os.path.isdir(os.path.join(setting_path, d))
        ]
        for mf_dir in mf_dirs:
            mf = extract_mf_from_dirname(mf_dir)
            if mf is None:
                continue

            mf_path = os.path.join(setting_path, mf_dir)

            ds_dirs = [
                d for d in os.listdir(mf_path)
                if os.path.isdir(os.path.join(mf_path, d))
            ]
            for ds_dir in ds_dirs:
                ds = normalize_dataset_name(ds_dir)
                train_log = os.path.join(mf_path, ds_dir, "train.log")

                metrics = parse_train_log_best_metrics(train_log)
                if metrics is None:
                    if miss_print_budget > 0:
                        print(f"[WARN] no best_ndcg5 line parsed: {setting_key}/{mf_dir}/{ds_dir}/train.log")
                        miss_print_budget -= 1
                    continue

                mf_data[mf][setting_key][ds] = metrics
                all_datasets_by_mf[mf].add(ds)

    if not mf_data:
        raise RuntimeError("No parsable train.log results were found.")

    wb = Workbook()
    wb.remove(wb.active)

    preferred = ["arxiv", "docvqa", "infovqa", "tabfquad", "tatdqa", "shift", "ai", "energy", "gov", "health"]

    for mf, setting_map in sorted(mf_data.items(), key=lambda x: float(x[0])):
        ds_set = set(all_datasets_by_mf[mf])
        ordered = [d for d in preferred if d in ds_set] + sorted([d for d in ds_set if d not in preferred])

        ws = wb.create_sheet(title=f"mf{mf}"[:31])
        write_sheet(ws, setting_map, ordered)

    wb.save(out_xlsx)
    print(f"[OK] Saved Excel to: {out_xlsx}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("Usage: python summary_results.py <root_dir> <out_xlsx_path>")
        raise SystemExit(1)
    build_excel_from_trainlogs(sys.argv[1], sys.argv[2])