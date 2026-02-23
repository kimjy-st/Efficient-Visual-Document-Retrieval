import os
import re
import json
from collections import defaultdict
from statistics import mean

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# -------------------------
# 1) train.log 마지막 summary 파서
# -------------------------
# 예: [2026-02-16 00:42:39,006][INFO] {"summary/best_recall":..., "summary/best_ndcg5":..., ...}
RE_SUMMARY_JSON = re.compile(r'(\{.*"summary/best_ndcg5".*\})\s*$')

def parse_last_summary_from_trainlog(train_log_path: str):
    """
    train.log를 뒤에서부터 훑어서 summary/best_ndcg5 JSON을 찾고,
    best_ndcg5의 epoch, NDCG@5, Recall@1 을 반환.
    없으면 None 반환.

    return:
      {"epoch": int, "N@5": float, "R@1": float}  # (이미 *100 적용 후 소수1자리 반올림)
    """
    try:
        with open(train_log_path, "r", encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()
    except Exception:
        return None

    for ln in reversed(lines):
        ln = ln.strip()
        m = RE_SUMMARY_JSON.search(ln)
        if not m:
            continue
        try:
            obj = json.loads(m.group(1))
        except json.JSONDecodeError:
            continue

        best = obj.get("summary/best_ndcg5", None)
        if not isinstance(best, dict):
            continue

        try:
            n5 = round(float(best.get("NDCG@5")) * 100.0, 1)
            r1 = round(float(best.get("Recall@1")) * 100.0, 1)
            return { "N@5": n5, "R@1": r1}
        except Exception:
            continue

    return None


# -------------------------
# 2) mf 추출 (폴더명 mf5/mf10...)
# -------------------------
RE_MF_DIR = re.compile(r"^mf(\d+)$", re.IGNORECASE)

def extract_mf_from_dirname(name: str):
    m = RE_MF_DIR.match(name.strip())
    if not m:
        return None
    return int(m.group(1))


# -------------------------
# 3) 엑셀 유틸 (너가 준 스타일 유지)
# -------------------------
def autosize_columns(ws):
    max_width = 28
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            s = str(cell.value)
            max_len = max(max_len, len(s))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

def style_header(ws, header_row=1):
    bold = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[header_row]:
        cell.font = bold
        cell.alignment = align

def add_table(ws, last_row, last_col):
    ref = f"A1:{get_column_letter(last_col)}{last_row}"
    tab = Table(displayName=f"Table_{ws.title}".replace(".", "_"), ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

def write_mf_sheet(ws, dataset_to_metrics, datasets_ordered):
    """
    한 mf 시트 작성.
    dataset_to_metrics: {ds: {"epoch": int, "N@5": float, "R@1": float}}
    """
    # 헤더: metric | (ds별 N@5, R@1) | average N@5, R@1
    header = ["metric"]
    for ds in datasets_ordered:
        header += [f"{ds}_N@5", f"{ds}_R@1"]
    header += ["average_N@5", "average_R@1"]
    ws.append(header)

    # 값 row: "best_ndcg5" 한 줄만 (너가 원하는 표 형태에 맞춤)
    row = ["best_ndcg5"]

    n5_list, r1_list = [], []

    for ds in datasets_ordered:
        m = dataset_to_metrics.get(ds)
        if m is None:
            row += [None, None]
        else:
            n5, r1 = m["N@5"], m["R@1"]
            row += [n5, r1]
            n5_list.append(n5)
            r1_list.append(r1)

    row += [
        round(mean(n5_list), 1) if n5_list else None,
        round(mean(r1_list), 1) if r1_list else None,
    ]
    ws.append(row)

    # 스타일/포맷
    style_header(ws, 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # 숫자 포맷: 전부 소수 1자리로 보여주기
    last_row = ws.max_row
    last_col = ws.max_column
    for r in range(2, last_row + 1):
        for c in range(2, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            cell.number_format = "0.0"

    autosize_columns(ws)
    add_table(ws, last_row, last_col)


# -------------------------
# 4) 메인: root/mf*/dataset/train.log 읽어서 mf별 시트 생성
# -------------------------
def build_mf_excel_from_trainlogs(root_dir: str, out_xlsx: str):
    root_dir = os.path.abspath(root_dir)
    os.makedirs(os.path.dirname(os.path.abspath(out_xlsx)), exist_ok=True)

    # mf 폴더 수집
    mf_dirs = []
    for name in os.listdir(root_dir):
        p = os.path.join(root_dir, name)
        if not os.path.isdir(p) or name.startswith("."):
            continue
        mf = extract_mf_from_dirname(name)
        if mf is not None:
            mf_dirs.append((mf, p))

    if not mf_dirs:
        raise RuntimeError(f"No mf* directories found under: {root_dir}")

    # mf -> {dataset: metrics}
    mf_data = defaultdict(dict)
    all_datasets_by_mf = defaultdict(set)

    for mf, mf_path in sorted(mf_dirs, key=lambda x: x[0]):
        # mf_path 아래 dataset 폴더들
        for ds_name in os.listdir(mf_path):
            ds_path = os.path.join(mf_path, ds_name)
            if not os.path.isdir(ds_path) or ds_name.startswith("."):
                continue

            ds = ds_name
            log_path = os.path.join(ds_path, "train.log")
            if not os.path.exists(log_path):
                continue

            metrics = parse_last_summary_from_trainlog(log_path)
            if metrics is None:
                continue

            mf_data[mf][ds] = metrics
            all_datasets_by_mf[mf].add(ds)

    if not mf_data:
        raise RuntimeError("No valid train.log with parsable summary/best_ndcg5 were found.")

    wb = Workbook()
    wb.remove(wb.active)

    # 네가 원하는 표 컬럼 순서 (있으면 이 순서 우선)
    preferred = ["arxiv", "docvqa", "infovqa", "tabfquad", "tatdqa", "shift", "ai", "energy", "gov", "health"]

    for mf in sorted(mf_data.keys()):
        ds_set = set(all_datasets_by_mf[mf])
        ordered = [d for d in preferred if d in ds_set] + sorted([d for d in ds_set if d not in preferred])
        print(ordered)
        ws = wb.create_sheet(title=f"mf{mf}"[:31])
        write_mf_sheet(ws, mf_data[mf], ordered)

    wb.save(out_xlsx)
    print(f"[OK] Saved Excel to: {out_xlsx}")


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("Usage: python trainlog2xlsx.py <root_dir> <out_xlsx_path>")
        raise SystemExit(1)

    build_mf_excel_from_trainlogs(sys.argv[1], sys.argv[2])